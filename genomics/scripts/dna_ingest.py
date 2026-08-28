"""Ingest raw genotipi (txt 23andMe) → data/dna/genotypes.db. Deterministico e idempotente."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import dna_common

NO_CALLS = {"--", ""}


def _clean_num(v) -> str:
    """Normalizza numeri da xlsx: 1.0 → "1", 2000.0 → "2000"."""
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def _rows_txt(path: Path):
    for line in path.read_text().splitlines():
        if not line or line.startswith("#"):
            continue
        parts = line.split("\t")
        if len(parts) != 4:
            continue
        rsid, chrom, pos, gt = parts
        yield rsid, chrom, int(pos), gt.strip()


def _rows_vcf(path: Path):
    for line in path.read_text().splitlines():
        if not line or line.startswith("#"):
            continue
        f = line.split("\t")
        if len(f) < 10 or not f[2].startswith("rs"):
            continue
        chrom, pos, rsid, ref, alt = f[0], int(f[1]), f[2], f[3], f[4]
        gt_field = f[9].split(":")[0].replace("|", "/")
        alleles = {"0": ref, "1": alt.split(",")[0]}
        try:
            a, b = gt_field.split("/")
            gt = alleles[a] + alleles[b]
        except (ValueError, KeyError):
            gt = "--"
        yield rsid, chrom, pos, gt


def _rows_xlsx(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("Per i file xlsx serve openpyxl: python3.12 -m pip install openpyxl")
    ws = load_workbook(path, read_only=True).active
    want = ["rsid", "chromosome", "position", "genotype"]
    idx = None
    for row in ws.iter_rows(max_row=50, values_only=True):
        # Normalizza: rimuovi "#" e spazi, poi lowercase
        low = [str(c).strip().lstrip("#").strip().lower() if c is not None else "" for c in row]
        if all(w in low for w in want):
            idx = {w: low.index(w) for w in want}
            header_found = row
            break
    if idx is None:
        raise SystemExit(f"{path.name}: intestazione dati non trovata (colonne attese: {want})")
    started = False
    for row in ws.iter_rows(values_only=True):
        if not started:
            started = row == header_found
            continue
        if row[idx["rsid"]] is None:
            continue
        chrom = _clean_num(row[idx["chromosome"]])
        pos = int(float(row[idx["position"]]))
        yield (str(row[idx["rsid"]]).strip(), chrom,
               pos, str(row[idx["genotype"]]).strip())


READERS = {".txt": _rows_txt, ".vcf": _rows_vcf, ".xlsx": _rows_xlsx}


def ingest(path: Path, build: str = "GRCh37") -> dict:
    path = Path(path)
    con = dna_common.connect()
    digest = dna_common.file_sha256(path)
    if dna_common.get_meta(con, f"ingested:{digest}"):
        return {"status": "already_ingested", "inserted": 0, "skipped_nocall": 0}
    reader = READERS.get(path.suffix.lower())
    if reader is None:
        raise SystemExit(f"Formato non supportato: {path.suffix} (attesi: {sorted(READERS)})")
    inserted = skipped = 0
    for rsid, chrom, pos, gt in reader(path):
        if gt in NO_CALLS:
            skipped += 1
            continue
        con.execute(
            "INSERT INTO genotypes(rsid,chrom,pos,genotype,source_file,build) "
            "VALUES(?,?,?,?,?,?) ON CONFLICT(rsid) DO UPDATE SET "
            "genotype=excluded.genotype, source_file=excluded.source_file",
            (rsid, chrom, pos, gt, path.name, build))
        inserted += 1
    dna_common.set_meta(con, "build", build)
    dna_common.set_meta(con, f"ingested:{digest}", path.name)
    con.commit()
    return {"status": "ok", "inserted": inserted, "skipped_nocall": skipped}


def main() -> None:
    p = argparse.ArgumentParser(description="Ingest raw DNA nel repo dati")
    p.add_argument("file")
    p.add_argument("--build", default="GRCh37")
    args = p.parse_args()
    res = ingest(Path(args.file), args.build)
    print(res)


if __name__ == "__main__":
    main()
